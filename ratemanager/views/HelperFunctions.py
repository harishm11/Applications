import os
import re
from django.shortcuts import redirect
from django.contrib import messages
import pandas as pd
from datetime import datetime
from django.apps import apps
from ratemanager.forms import searchCriteriaForm
from ratemanager.models.ratebookmetadata import RatebookMetadata
from ratemanager.models.ratebooktemplate import RatebookTemplate
from ratemanager.models.ratingfactors import RatingFactors
from ratemanager.models.ratingexhibits import RatingExhibits
from ratemanager.models.ratingvariables import RatingVariables
from myproj.settings import BASE_DIR
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.utils.html import format_html
from copy import deepcopy
import openpyxl
import sqlalchemy as sa
from myproj.settings import DATABASES
from ratemanager.views.configs import ENVIRONMENT_HIERARCHY
from myproj.messages import RATE_MANAGER
from difflib import get_close_matches
import ratemanager.views.configs as configs

coverage = apps.get_model('systemtables', 'coverage')

SIDEBAR_OPTIONS = ['rates', "template", "createRB", "viewRB", "updateRB",
                   "viewRBbyDate", "viewRBbyVersion",
                   "viewTemplateOptions", 'EditCoverages']
pd.options.mode.copy_on_write = True

list_of_known_covs = [
    'BI', 'PD', 'UMBI', 'UMPDC-1', 'UMPDC-2',
    'MED', 'COMP', 'COLL', 'LOSS OF USE', 'RENT', 'TOW'
]
separators_list = ['-', u'\u2013', '&', '/', 'to', '\\']
# required for transform
exhibits_to_handle_manually = [
    'DeductiblesbySymbol', 'DriverTrainingDiscount', 'UMPDC2BaseRatesbyDeductible',
    'CollisionRatesTrailers', 'CCDRatesTrailers', 'CollisionRatesCampers', 'CCDRatesCampers'
]
example_values_of_last_ratingVar = {
    "Affinity Group": ("GROUP I", "GROUP II"),
}

excludeList = ["Ratebook Details", "Rate Details", "RatebookDetails",
               "DELETED_EXHIBITS", 'RENAMED_EXHIBITS',
               'PointsAssignmentsDPS']


def find_separator(s):
    for x in separators_list:
        if x in s:
            return True, x
    return None, None


def check_all_numeric(range, separator):
    ls = range.split(separator)
    if all([x.replace(' ', '').isnumeric() for x in ls]):
        return True


def check_range(pds):
    if not isinstance(pds, pd.Series):
        pds = pds.to_series()

    def found_separators(s):
        if isinstance(s, str):
            has_separator, separator = find_separator(s)
            if any([has_separator and check_all_numeric(s, separator)
                    for x in separators_list]):
                return True
        return False
    if pds.apply(found_separators).any():
        return True
    return False


def create_rate_vars_cols(df_in, var_list):
    rename_col_names = []
    df_in = df_in[df_in.columns.sort_values().to_list()]
    for i in range(1, len(var_list)+1):
        rename_col_names.append((var_list[i-1], 'RatingVarValue'+str(i)))
    rename_col_names = dict(rename_col_names)
    for i in range(1, len(var_list)+1):
        df_in['RatingVarName'+str(i)] = var_list[i-1]
    df_in.rename(columns=rename_col_names, inplace=True)
    return df_in


def categoriseTransformType(df, sheet_name):
    # check for any coverages in the columns of dataframes
    # (cov as a subset of col_name and vice versa)
    def find_ratevars(possible_list):
        '''
        possible_list: list of df column names
        '''
        blocklist = ['symbol', 'band']
        for i in list_of_known_covs:
            for index, j in enumerate(possible_list):
                i = i.lower()
                j = j.lower()
                if (((i == j) or (i in j) or (j in i)) or 'factor' in j) \
                        and not any([x in j for x in blocklist]) \
                        and ('unnamed:' not in j):
                    ratevars = list(possible_list)[:index]
                    if 'Coverage' in ratevars:
                        ratevars.remove('Coverage')
                    covs = list(possible_list)[index:]
                    return ratevars, covs
        return [], []

    tabStruct = dict()
    ratevars, covs = find_ratevars(df.columns)
    tabStruct['ratevars'] = ratevars
    tabStruct['covs'] = covs

    # assumption if covs as rows there is a column called coverage
    if any(['coverage' == x.lower() for x in list(df.columns)]):
        tabStruct['category'] = 'covs as rows'

    elif any(['factor' in x.lower() for x in list(df.columns)]):
        tabStruct['category'] = 'no covs but factor in cols'

    elif (ratevars or covs):
        tabStruct['category'] = 'covs as cols'

    elif not covs:
        tabStruct['category'] = 'neither covs nor factor in cols'
        tabStruct['ratevars'] = df.columns.to_list()

    tabStruct['hasRange'] = []

    # check for range in columns
    for i in df.columns:
        if check_range(df[i]):
            tabStruct['hasRange'].append(i)
    return tabStruct


def handle_manually(tabStruct, sheetname, df_in):
    """
    This function takes in a tabStruct, sheetname, and a dataframe (df_in) as input.
    It then performs various operations on the dataframe based on the sheetname and returns the modified dataframe.

    Parameters:
    tabStruct (dict): A dictionary containing various parameters for the function.
    sheetname (str): The name of the sheet to be processed.
    df_in (pandas.DataFrame): The input dataframe to be processed.

    Returns:
    pandas.DataFrame: The modified dataframe after performing various operations based on the sheetname.
    """
    if sheetname == 'DeductiblesbySymbol':
        df_in = df_in.melt(id_vars=tabStruct['ratevars'])
        df_in.rename(columns={'variable': 'Coverage',
                     'value': 'Factor'}, inplace=True)

        df_in['Symbol1'] = pd.NA
        df_in['Symbol2'] = pd.NA
        for index, val in df_in['Coverage'].items():
            cov = r = None
            has_separator, separator = find_separator(val)
            if has_separator:
                cov, r = val.split()
            if separator == '-':
                df_in.loc[index, 'Coverage'] = cov
                df_in.loc[index, 'Symbol1'], \
                    df_in.loc[index, 'Symbol2'] = r.split(separator)
            if separator == '&':
                df_in.loc[index, 'Coverage'] = cov
                a, b = r.split(separator)
                df_in.loc[index, 'Symbol1'] = a
                df_in.loc[int(df_in.index[-1])+1] = df_in.loc[index]
                df_in.loc[int(df_in.index[-1]), 'Symbol1'] = b
        tabStruct['ratevars'].extend(['Symbol1', 'Symbol2'])
        df_in = create_rate_vars_cols(df_in, tabStruct['ratevars'])
        df_in['Exhibit'] = sheetname
        return df_in
    if sheetname == 'DriverTrainingDiscount':
        newdf = df_in.melt(id_vars=tabStruct['ratevars'])
        newdf.rename(columns={'variable': 'Coverage',
                     'value': 'Factor'}, inplace=True)
        for index, val in newdf['Description'].items():
            if any([x.isnumeric() for x in val.split()]):
                cleanedVal = [x for x in val.split() if x.isnumeric()]
                if len(cleanedVal) >= 2:
                    newdf.loc[index, 'Description'] = cleanedVal[0]
                    for i in range(int(cleanedVal[0])+1, int(cleanedVal[1])+1):
                        newdf.loc[int(newdf.index[-1])+1] = newdf.loc[index]
                        newdf.loc[int(newdf.index[-1]), 'Description'] = str(i)
                else:
                    newdf.loc[index, 'Description'] = cleanedVal[0]
        newdf.sort_values(by=['Description', 'Coverage'], inplace=True)
        newdf = create_rate_vars_cols(newdf, tabStruct['ratevars'])
        newdf['Exhibit'] = sheetname
        newdf.loc[newdf['RatingVarValue1'] ==
                  'Senior Defensive', 'Exhibit'] = 'SeniorDefensive'
        return newdf

    if sheetname == 'UMPDC2BaseRatesbyDeductible':
        newdf = df_in.melt(id_vars='Deductible')
        newdf.rename(columns={'variable': 'AffinityGroup',
                     'value': 'Factor'}, inplace=True)
        newdf = create_rate_vars_cols(newdf, ['Deductible', 'AffinityGroup'])
        newdf['Exhibit'] = sheetname
        newdf['Coverage'] = 'UMPD'
        return newdf

    if sheetname in ['CollisionRatesTrailers', 'CCDRatesTrailers', 'CollisionRatesCampers', 'CCDRatesCampers']:
        df_in = df_in.melt(id_vars=df_in.columns[0:2])
        df_in.rename(columns={'variable': 'Deductible',
                     'value': 'Factor'}, inplace=True)

        def clean(string):
            if isinstance(string, str):
                return ''.join([x for x in string if x not in [' ', '$', '\n', u'\24']])
        df_in.apply(clean)
        df_in['Amount1'] = pd.NA
        df_in['Amount2'] = pd.NA
        for index, val in df_in[df_in.columns[0]].items():
            val = ''.join([x for x in val if x not in [' ', '$']])
            r = None
            has_separator, separator = find_separator(val)
            if has_separator:
                r1, r2 = val.split(separator)
                if separator == '-' or separator == u'\u2013':
                    df_in.loc[index, 'Amount1'], \
                        df_in.loc[index, 'Amount2'] = r1, r2
                if separator == '&':
                    df_in.loc[index, 'Amount1'] = r1
        ratevars = ['Amount1', 'Amount2', 'Deductible']
        ratevars.append(df_in.columns[1])
        df_in.drop(df_in.columns[0], axis=1, inplace=True)
        df_in = create_rate_vars_cols(df_in, ratevars)
        df_in['Exhibit'] = sheetname
        df_in['Coverage'] = 'COLL'
        return df_in

    if sheetname.startswith('FrequencyandSeverity'):
        keySubstrings = ['FREQ BAND', 'SEV BAND']

        def filterAndTransform(keySubstring, df):
            filtered_cols = ['ZIP CODE'] + \
                [col for col in df.columns if keySubstring in col]
            filtered = df[filtered_cols]
            new_cols = [col.replace(keySubstring, '').strip()
                        for col in filtered.columns]
            filtered.columns = new_cols
            filtered = filtered.melt(
                id_vars=['ZIP CODE'], value_vars=filtered.columns[1:])
            filtered.rename(
                columns={'variable': 'Coverage', 'value': keySubstring}, inplace=True)
            return filtered

        to_joins = [filterAndTransform(i, df=df_in) for i in keySubstrings]
        merged = pd.merge(to_joins[0], to_joins[1],
                          on=['ZIP CODE', 'Coverage'])
        merged['Exhibit'] = sheetname
        return create_rate_vars_cols(merged, ['ZIP CODE', 'FREQ BAND', 'SEV BAND'])

    else:
        return df_in


def get_table_category(sheet_name):
    nonFactorExhibits = ["severitybands", "pointsassignment"]
    if any([x.lower() in sheet_name.lower() for x in nonFactorExhibits]):
        return "Non-Factor Table"
    else:
        return "Factor Table"


def transform(sheet_name, df):
    pd.options.mode.copy_on_write = True
    tabStruct = categoriseTransformType(df, sheet_name)
    newdf = pd.DataFrame()

    def find_cov_name(sheetname):
        for i in list_of_known_covs:
            if i.lower() in sheetname.lower():
                return i
        return 'All'

    def handle_ranges(df_in):
        if tabStruct['hasRange']:
            for i in tabStruct['hasRange']:
                df_in[str(i)+'1'] = pd.NA
                df_in[str(i)+'2'] = pd.NA
                for index, val in df_in[i].items():
                    has_separator, separator = find_separator(val)
                    if has_separator:
                        df_in.loc[index, str(
                            i)+'1'], df_in.loc[index, str(i)+'2'] = val.split(separator)
                    else:
                        df_in.loc[index, str(i)+'1'] = val
                df_in.drop([i], axis=1, inplace=True)
                tabStruct['ratevars'].remove(i)
                tabStruct['ratevars'].extend([str(i)+'1', str(i)+'2'])

    if sheet_name in exhibits_to_handle_manually or 'FrequencyandSeverity' in sheet_name:
        newdf = handle_manually(tabStruct, sheet_name, df)
    else:
        if tabStruct['category'] == 'covs as cols':
            df = df.melt(id_vars=tabStruct['ratevars'])
            df.rename(columns={'variable': 'Coverage',
                      'value': 'Factor'}, inplace=True)
            handle_ranges(df)
            newdf = create_rate_vars_cols(df, tabStruct['ratevars'])
            newdf['Exhibit'] = sheet_name
        elif tabStruct['category'] == 'covs as rows':
            df.rename(columns={'Factor Amt.': 'Factor'}, inplace=True)
            handle_ranges(df)
            newdf = create_rate_vars_cols(df, tabStruct['ratevars'])
            newdf['Exhibit'] = sheet_name
        elif tabStruct['category'] == 'no covs but factor in cols':
            df.rename(columns={'Factor Amt.': 'Factor'}, inplace=True)
            handle_ranges(df)
            ratevars = []
            for i in df.columns:
                if 'factor' not in i.lower():
                    ratevars.append(i)
            newdf = create_rate_vars_cols(df, ratevars)
            newdf['Coverage'] = find_cov_name(sheet_name)
            newdf['Exhibit'] = sheet_name
        elif tabStruct['category'] == 'neither covs nor factor in cols':
            newdf = create_rate_vars_cols(df, tabStruct['ratevars'])
            newdf['Exhibit'] = sheet_name
            newdf['Coverage'] = find_cov_name(sheet_name)
            return newdf
        else:
            return df

    return newdf


def transformRB(xl_url=None, df_sheets=None):
    """Function to transform the excel file to required form
    to be able to load into postgres database."""

    msgs = []
    if xl_url:
        df_sheets = pd.read_excel(xl_url, sheet_name=None)

    df_dict = dict()
    df_out = pd.DataFrame(columns=["Exhibit", "Coverage", "Factor"])

    for sheet_name, df in df_sheets.items():
        if sheet_name not in excludeList:
            df_dict[sheet_name] = transform(sheet_name, df)
            df_dict[sheet_name]['TableCategory'] = get_table_category(
                sheet_name)
            df_dict[sheet_name] = df_dict[sheet_name].astype(object)
            try:
                df_out = df_out.merge(df_dict[sheet_name], how='outer')
            except Exception as err:
                print("Unable to transform {} table due to {}".format(
                    sheet_name, err))

    numRatingVars = configs.NO_OF_RATING_VARIABLES
    for i in range(1, numRatingVars+1):
        if 'RatingVarName' + str(i) not in df_out.columns:
            df_out['RatingVarName' + str(i)] = None
            df_out['RatingVarValue' + str(i)] = None

    df_out = df_out[df_out.columns.sort_values().to_list()]
    df_out = df_out.astype(str).astype(object).where(pd.notnull(df_out), None)

    for i in set(df_sheets.keys()) - set(df_out["Exhibit"].unique()):
        if i not in excludeList:
            msgs.append("Unable to transform {} table.".format(i))
    # map_covs_and_vars(df_out)
    # df_out.to_excel('./uploads/check_transform.xlsx')
    return df_out, msgs


def convertDates(x):
    """converts date from mm-dd-yy to django acceptable format"""
    if ':' in x:
        return datetime.strptime(x, "%Y-%d-%m %H:%M:%S").strftime("%Y-%m-%d")
    elif "/" in x:
        return datetime.strptime(x, "%m/%d/%Y").strftime("%Y-%m-%d")
    elif "-" in x:
        return datetime.strptime(x, "%m-%d-%Y").strftime("%Y-%m-%d")


def applyDateConversion(rate_details):
    ''' apply date conversion from m/d/y to django required format'''
    for key in rate_details:
        if "Date" in key:
            rate_details[key] = convertDates(rate_details[key])
    return rate_details


def fetchForeignFields(rate_details):
    ''' This function fetches the foreign key field objects in ratebook details while creating a ratebook. '''
    # get models from other apps
    uwCompany = apps.get_model("systemtables", "uwcompany")
    state = apps.get_model("systemtables", "state")
    carrier = apps.get_model("systemtables", "carrier")
    lineOfBusiness = apps.get_model("systemtables", "lineofbusiness")
    policyType = apps.get_model("systemtables", "policytype")
    policySubType = apps.get_model("systemtables", "policysubtype")
    productCode = apps.get_model("systemtables", "productcode")
    rate_details["Carrier"] = carrier.objects.get(
        CarrierName=rate_details["Carrier"])
    rate_details["State"] = state.objects.get(StateName=rate_details["State"])
    rate_details["LineofBusiness"] = lineOfBusiness.objects.get(
        LobName=rate_details["LineofBusiness"]
    )
    rate_details["UWCompany"] = uwCompany.objects.get(
        CompanyName=rate_details["UWCompany"]
    )
    rate_details["PolicyType"] = policyType.objects.get(
        PolicyTypeName=rate_details["PolicyType"]
    )
    rate_details["PolicySubType"] = policySubType.objects.get(
        PolicySubTypeName=rate_details["PolicySubType"]
    )
    rate_details["ProductCode"] = productCode.objects.get(
        ProductCd=rate_details["ProductCode"]
    )
    return rate_details


def generateRatebookID():
    ''' This function generates a Ratebook ID of the format "RB00XX" by incrementing the existing largest ID '''
    LargestIDObj = RatebookMetadata.objects.filter().order_by("-RatebookID").first()
    LargestID = LargestIDObj.RatebookID if LargestIDObj is not None else None
    if LargestID is not None:
        head = LargestID.rstrip("0123456789")
        tail = LargestID[len(head):]
        newrb = head + str(int(tail) + 1).zfill(4)
        return newrb
    else:
        return "RB0001"


def loadtoRatingFactors(df):
    ''' This function loads a Dataframe to RatingFactors Model '''

    db_url = 'postgresql://' + DATABASES['default']['USER'] + ':' \
        + DATABASES['default']['PASSWORD'] + '@'\
        + DATABASES['default']['HOST'] + ':'\
        + DATABASES['default']['PORT'] + '/'\
        + DATABASES['default']['NAME']
    print('Loading to {}'.format(db_url))
    engine = sa.create_engine(db_url)
    df.to_sql('ratemanager_ratingfactors', engine, chunksize=1000,
              method='multi', if_exists='append', index=False)


def findRBDetails(df):
    """ Function to find Ratebook Details Tab
        returns ratebook details sheet name """

    rateDetailsSheet = None
    for sheet in list(df.keys()):
        iLower = sheet.lower()
        if ("rate" in iLower or "book" in iLower) and "details" in iLower:
            rateDetailsSheet = sheet
    return rateDetailsSheet


def uploadFile(request):
    ''' Function to get the uploaded file from request and save it.
        returns the file path url. '''
    # save and get uploaded file path
    upfile = request.FILES.get("file")
    # print(BASE_DIR)
    root = os.path.join(BASE_DIR, "uploads")
    path = os.path.abspath(os.path.join(
        root, str(upfile.name.replace(" ", ""))))
    # print(path)
    fileexists = False
    if not fileexists:
        filstg = FileSystemStorage(base_url=str(BASE_DIR))
        upldfl = filstg.save(path, upfile)
        # upldfl_url = filstg.url(upldfl)
        upldfl_url = os.path.join(filstg.location, upldfl)
    return upldfl_url


def fetchRatebookSpecificVersion(rbID, rbVersion):

    ActivationDate = RatebookMetadata.objects.get(
        RatebookID=rbID,
        RatebookVersion=rbVersion
    ).ActivationDate

    current_version = fetchRatebookbyDate(rbID, ActivationDate)

    return current_version


def fetchRatebookbyDate(rbID, qDate):
    '''
    Select RATEBOOKID from Ratemanager_ratingFactors
    where NB Effective <= current date
    and RN Effective Dt < = current Date
    and (( NB expiry and rn Expiry  is null)
    or (NB expiry and RN expiry date > current date) )
    and activation date <= current date
    '''
    filteredQueryResults = RatingFactors.objects.filter(RatebookID=rbID).filter(
        Q(NewBusinessEffectiveDate__lte=qDate)
        & Q(RenewalEffectiveDate__lt=qDate)
        & (
            (Q(NewBusinessExpiryDate__isnull=True)
             & Q(RenewalExpiryDate__isnull=True))
            | (Q(NewBusinessExpiryDate__gt=qDate) & Q(RenewalExpiryDate__gt=qDate))
        )
        & (Q(ActivationDate__lte=qDate))
    )
    return filteredQueryResults


def extractRatebookDetails(inputDataFrame):
    """ Checks all the sheets for RB Details and returns it as DataFrame and HTML Table """
    df = inputDataFrame
    msgs = []
    sheet_name = findRBDetails(df)

    if sheet_name is not None:
        msgs.append("Found Ratebook Details")
        df = df[sheet_name]
        df[0] = df[0].str.replace(" ", "")
        df_view = pd.Series(index=list(
            df[0]), data=list(df[1]), name="Details")
        df_view.astype(str).replace("nan", None, inplace=True)
        rbDetailsTable = format_html(
            df_view.to_frame().to_html(
                justify="left", classes=["table", "table-bordered"]
            )
        )
    else:
        msgs.append("Could not find Ratebook Details Sheet\
                     in the uploaded Excel file please check again.")
    return {'details_html': rbDetailsTable,
            'details_df': df_view,
            'msgs': msgs}


def fetchRBLatestVersion(rbID):
    '''
    Loads latest version from RatingFactors of given RBID as a dataframe.
    '''
    latestVersion = (
        RatebookMetadata.objects.filter(RatebookID=rbID)
        .order_by("-RatebookVersion")
        .first()
    ).RatebookVersion
    oldRB = fetchRatebookSpecificVersion(
        rbID=rbID,
        rbVersion=latestVersion,
    )
    return oldRB


def dataframe_difference(old_df, new_df):
    """Find rows which are added, deleted or modified in the new version of DataFrame."""
    def convertToObject(df):
        ''' convert all columns to object dtype except factor '''
        for col in list(df.columns):
            df[col] = df[col].astype(object)
        return df
    stats = {}
    stats['isEmpty'] = True
    if new_df.empty:
        return {
            'modified': None,
            'added': None,
            'deleted': None,
        }, stats

    old_df = convertToObject(old_df)
    new_df = convertToObject(new_df)

    comparison_df = old_df.merge(
        new_df,
        indicator=True,
        how='outer'
    )

    comparison_df = comparison_df.astype(
        object).where(pd.notnull(comparison_df), None)

    deleted_old = comparison_df.loc[comparison_df._merge == 'left_only']
    added_new = comparison_df.loc[comparison_df._merge == 'right_only']

    modified = deleted_old.merge(
        added_new,
        how='inner',
        on=[x for x in list(new_df.columns) if x not in ('Factor', '_merge')]
    )
    modified.drop(['_merge_x', '_merge_y'], inplace=True, axis=1)
    modified.rename(columns={'Factor_x': 'Factor_Old',
                             'Factor_y': 'Factor_New'},
                    inplace=True)
    modified = modified[modified.columns.sort_values().to_list()]

    modified = modified.astype(object).where(pd.notnull(modified), None)

    added_deleted = comparison_df.drop_duplicates(
        subset=[x for x in list(new_df.columns)
                if x not in ('Factor', '_merge')],
        ignore_index=True,
        keep=False
    )

    added = added_deleted.loc[added_deleted._merge ==
                              'right_only'].drop('_merge', axis=1)
    deleted = added_deleted.loc[added_deleted._merge ==
                                'left_only'].drop('_merge', axis=1)

    added = added.astype(object).where(pd.notnull(added), None)
    deleted = deleted.astype(object).where(pd.notnull(deleted), None)

    stats['Number of rate revisions'] = len(modified)
    stats['Number of added factors'] = len(added)
    stats['Number of deleted factors'] = len(deleted)
    stats['changed_exhibits'] = []
    stats['changed_exhibits'].extend(added['Exhibit'].unique())
    stats['changed_exhibits'].extend(deleted['Exhibit'].unique())
    stats['changed_exhibits'].extend(modified['Exhibit'].unique())
    stats['changed_exhibits'] = set(stats['changed_exhibits'])
    stats['isEmpty'] = False if len(
        modified)+len(added)+len(deleted) > 0 else True

    return {'modified': modified,
            'added': added,
            'deleted': deleted,
            }, stats


def generate_html_diff(changes):
    ''' Generate Details about the changes in between the 2 dataframes '''
    updates = deepcopy(changes)
    updates['modified']['Factor'] = updates['modified']['Factor_Old'].astype(str) +\
        ' -> ' + updates['modified']['Factor_New'].astype(str)
    updates['modified'].drop(
        ['Factor_Old', 'Factor_New'], axis=1, inplace=True)
    updates['modified'] = updates['modified'][updates['modified'].columns.sort_values().to_list()]
    updates['added'] = updates['added'][updates['modified'].columns.sort_values().to_list()]
    updates['deleted'] = updates['deleted'][updates['modified'].columns.sort_values().to_list()]
    updates['modified'] = updates['modified'].astype(str).style.applymap(
        lambda x: "background-color: lightblue" if '->' in x else "background-color: white")
    updates['added'] = updates['added'].astype(str).style.applymap(
        lambda x: "background-color: lightgreen")
    updates['deleted'] = updates['deleted'].astype(str).style.applymap(
        lambda x: "background-color: lightcoral")
    diff_df = updates['modified'].concat(updates['deleted'])
    diff_df = diff_df.concat(updates['added'])
    diff_df = diff_df.hide(axis='index')
    rbChangesTableHTML = diff_df.to_html(
        table_uuid='changes'
    )
    return rbChangesTableHTML


def convert2Df(QuerySet):
    ''' Convert QuerySet to Dataframe and
    Drops fields not needed for comparision of Rate Factors '''
    toDropList = ['id', 'RatebookID', 'Ratebook_id',
                  "RatebookVersion", "RecordStatus", "TableCategory"]
    for i in [f.name for f in RatingFactors._meta.get_fields(include_hidden=False)]:
        if 'Date' in i or 'Time' in i:
            toDropList.extend([i])
    df = pd.DataFrame.from_records(QuerySet.values())
    df = df.astype(str).astype(object).where(pd.notnull(df), None)
    if not df.empty:
        df.drop(columns=toDropList, axis=1, inplace=True)
    return df


def extractUpdateRBDetails(xl_url, withExhibitStatusHeader=False):
    '''
    It was to be used to extract the exhibit status from the excel file in each sheet.
    Not used/needed anymore. can be removed.
    '''
    wb = openpyxl.load_workbook(xl_url)
    rbDetailSheetName = None
    for i in wb.sheetnames:
        if 'rate' in i.lower() and 'details' in i.lower():
            rbDetailSheetName = i
    sheet = wb[rbDetailSheetName]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    df = pd.DataFrame(data)

    # split to multiple tables
    SheetTableList = []
    table = []

    for row in df.itertuples(index=False):
        if any(row._asdict().values()):
            table.append(list(row._asdict().values()))
        elif len(table) > 0 and all([not x for x in row._asdict().values()]):
            print('Found Table Ending')
            SheetTableList.append(
                deepcopy(pd.DataFrame(table).dropna(axis=1, how='all')))
            table.clear()
    if len(table) != 0:
        SheetTableList.append(
            deepcopy(pd.DataFrame(table).dropna(axis=1, how='all')))

    toDelete = SheetTableList[1]
    return {'rbDetails'
            'exhibitStatusSeries': None,
            'deletedExhibits': toDelete}


def getToExpireExhibits(xl_url):
    '''
    returns
        'deletedList': deletedList,
        'renamedDict': renamedDict,
        'toExpire': toExpire
    '''
    toExpire = []
    deletedList = []
    renamedDict = dict()
    sheets = pd.read_excel(xl_url, sheet_name=None)
    for sheet, df in sheets.items():
        if sheet == 'DELETED_EXHIBITS':
            deletedList = df['Deleted Exhibit Name'].tolist()
            toExpire.append(deletedList)
        if sheet == 'RENAMED_EXHIBITS':
            renamedDict = df.to_dict()
            toExpire.extend(renamedDict['Old Exhibit Name'].values())

    return {
        'deletedList': deletedList,
        'renamedDict': renamedDict,
        'toExpire': toExpire
    }


def inverseTransform(df):
    '''
    This function takes the UnPivot View of the Dataframe and converts it to
    Pivot View i.e. the original excel sheet view. This is used to export the
    data from database to excel file and download it and also to display the
    data in the frontend.
    '''
    def pivotFromat(df, index, columns, values='Factor'):
        idf = df.pivot(
            index=index,
            columns=columns,
            values=values
        )

        idf = idf.reset_index(allow_duplicates=False)

        idf.columns = list(idf.columns)

        rename_dict = dict()
        for i in df.columns:
            if 'RatingVar' in i:
                rename_dict[i] = df['RatingVarName'+i[-1]].unique()[0]

        idf.rename(columns=rename_dict, inplace=True)
        return idf

    df.dropna(how='all', axis=1, inplace=True)

    if 'Factor' not in df.columns:
        df['Factor'] = None
    # for example base rates minor covs like tables
    if not any([re.findall(r"\ARatingVar", x) for x in list(df.columns)]):
        return df[['Coverage', 'Factor']]

    varCount = sum([1 if re.match(r'\ARatingVar', x)
                   else 0 for x in list(df.columns)])/2
    varNames = [x for x in list(
        df.columns) if re.match(r'\ARatingVarValue', x)]

    # one Rating variable and many covs
    if varCount == 1:
        return pivotFromat(df, index='RatingVarValue1', columns='Coverage')

    if varCount >= 2:
        if df['Coverage'].unique()[0] == 'All':
            return pivotFromat(df, index=varNames, columns='Factor')
        if len(df['Coverage'].unique()) == 1:
            return pivotFromat(df, index=varNames[1:], columns=varNames[0])
        else:
            return pivotFromat(df, index=varNames, columns='Coverage')


def map_covs_and_vars(df):
    '''
    This function maps the Coverage and RatingVarName columns to their respective codes
    as per the system tables/policy processing system.

    Right now it is using the configs.py file to map the Coverage and RatingVarName columns
    '''
    ratevars = configs.ratevars
    ratevar_mappings = dict()
    uniq_ratevars = []
    for i in df.columns:
        if 'RatingVarName' in i:
            uniq_ratevars.extend(list(df[i].unique()))

    uniq_ratevars = [x for x in uniq_ratevars if isinstance(x, str)]

    for i in uniq_ratevars:
        matches = get_close_matches(i, ratevars, n=1)
        ratevar_mappings[i] = matches[0] if matches else i

    for i in df.columns:
        if 'RatingVarName' in i:
            df[i] = df[i].replace(ratevar_mappings)
    df['Coverage'] = df['Coverage'].replace(configs.cov_mapping)


def camel_case_split(str):
    ''' split camel case string to list of words '''
    return re.findall(r'[A-Z](?:[a-z]+|[A-Z]*(?=[A-Z]|$))', str)


def generateRatebookItemCode(ExhibitName):
    ''' generate ratebook item code from exhibit name '''
    # return ''.join(list(map(str, map(ord, ''.join(camel_case_split(ExhibitName))))))[:100]
    return 1234


def determine_all_ratevars(df, tabStruct):
    ''' if no ratevars are specified in the sheet, then determine all ratevars '''
    pass


def updateRatingVars(uploadURL, TemplateID):
    '''
    This function is called by update rating Exhibits to
    update rating variables in template and their Unique Tables.

    Only called when a new template is created from excel file.
    '''
    TempObj = RatebookTemplate.objects.get(id=TemplateID)
    ExhibitObj = TempObj.RatebookExhibit
    sheetname = ExhibitObj.Exhibit

    # check if sheetname exists in original excel file
    # will happen in case new exhibits are created from existing ones
    # if dosen't exist, return because nothing to do
    try:
        df = pd.read_excel(uploadURL, sheet_name=sheetname)
    except ValueError:
        return

    # update ratevars in template and their Unique Tables if not already exists
    if sheetname not in excludeList:
        # fetch the Table Structure it contains RatingVars and Coverage
        tabStruct = categoriseTransformType(df, sheetname)
        # if no ratevars are specified in the sheet, then determine all ratevars
        if tabStruct['category'] == 'neither covs nor factor in cols':
            determine_all_ratevars(df, tabStruct)
        # guess the datatypes of the columns using pandas builtin function
        guessed_dtypes_df = df.convert_dtypes()
        # iterate over the columns and create RatingVar objects if not already exists
        for _, col in enumerate(df.columns):
            if tabStruct['ratevars'] is not None and col in tabStruct['ratevars']:
                RatingVarObj, _ = RatingVariables.objects.get_or_create(
                    RatingVarName=col)
                RatingVarObj.DisplayName = col
                RatingVarObj.RatingVarType = guessed_dtypes_df.dtypes[col]
                RatingVarObj.save()
                # add the RatingVar to the current Template
                TempObj.ExhibitVariables.add(RatingVarObj)


def updateRatingExhibits(tdf, rbid, uploadURL):
    '''
    This function updates rating exhibits and rating variables in template and their Unique Tables
    Only called when a new template is created from excel file.

    Takes the transformed dataframe and ratebook id as input and excel file path as input,
    excel file path is used by the updateRatingVars function to update rating variables in the order
    maintaind in the excel file.
    '''
    # check if Template already exists for this Ratebook id and return if it does
    if RatebookTemplate.objects.filter(RatebookID=rbid).exists():
        return

    # get Exhibits list from the dataframe
    ExList = tdf["Exhibit"].unique()
    for i in ExList:
        TempRec = RatebookTemplate()
        TempRec.RatebookID = rbid.split('_')[0]  # remove version from rbid
        ExhibitObj, _ = RatingExhibits.objects.get_or_create(Exhibit=i, DisplayName=' '.join(camel_case_split(i)))
        TempRec.RatebookExhibit = ExhibitObj
        df = tdf[tdf["Exhibit"] == i]
        TempRec.save()
        for i in df['Coverage'].unique().tolist():
            cov, _ = coverage.objects.get_or_create(CoverageCode=configs.cov_mapping.get(i, i), CoverageName=i)
            # add the Coverage to the current Template
            TempRec.ExhibitCoverages.add(cov)

        updateRatingVars(uploadURL=uploadURL, TemplateID=TempRec.id)


def extractIdentityDetails(dictionary: dict) -> dict:
    ''' extract identity details ('Carrier', 'State', 'LineofBusiness',
    'UWCompany', 'PolicyType', 'PolicyType', 'PolicySubType', 'ProductCode')
    from dictionary input and return a dictionary of identity details excluding all other keys '''

    identityKeys = ('Carrier', 'State', 'LineofBusiness', 'UWCompany', 'PolicyType',
                    'PolicyType', 'PolicySubType', 'ProductCode')
    identityRateDetails = {key: dictionary.get(key) for key in identityKeys}
    if not any(identityRateDetails.values()):
        identityRateDetails = {key+'_id': dictionary.get(key+'_id') for key in identityKeys}
    return identityRateDetails


def clone_object(obj, attrs={}):
    '''
    recursively clone an object and all its related objects
    '''
    # we start by building a "flat" clone
    clone = obj._meta.model.objects.get(pk=obj.pk)
    clone.pk = None

    # if caller specified some attributes to be overridden,
    # use them
    for key, value in attrs.items():
        setattr(clone, key, value)

    # save the partial clone to have a valid ID assigned
    clone.save()

    # Scan field to further investigate relations
    fields = clone._meta.get_fields()
    for field in fields:

        # Manage M2M fields by replicating all related records
        # found on parent "obj" into "clone"
        if not field.auto_created and field.many_to_many:
            for row in getattr(obj, field.name).all():
                getattr(clone, field.name).add(row)

        # Manage 1-N and 1-1 relations by cloning child objects
        if field.auto_created and field.is_relation:
            if field.many_to_many:
                # do nothing
                pass
            else:
                # provide "clone" object to replace "obj"
                # on remote field
                attrs = {
                    field.remote_field.name: clone
                }
                children = field.related_model.objects.filter(
                    **{field.remote_field.name: obj})
                for child in children:
                    clone_object(child, attrs)
    return clone


def checkTemplateCreateButtonEnable(obj):
    if configs.ENVIRONMENT_HIERARCHY.get(obj.Environment) == 0:
        return False
    else:
        identDetails = extractIdentityDetails(obj.__dict__)
        sortedSearcResults = RatebookMetadata.objects.filter(**identDetails) \
            .order_by('-RatebookID', '-RatebookVersion').exclude(Environment='Draft')
        if obj.id == sortedSearcResults.first().id:
            return True


def searchCriteriaProcessor(request):
    ''' returns 'searchCriteriaForm': form,
                'searchResults': searchResults,
                'searchResultTableHeaders': searchResultTableHeaders,
                'ENVIRONMENT_HIERARCHY': ENVIRONMENT_HIERARCHY'''
    searchResults = None
    # Left overs 'id', 'Environment', 'isDeleted', 'onHold', 'retrofitReq', 'Carrier', 'State', 'LineofBusiness', 'UWCompany', 'PolicyType', 'PolicySubType', 'ProductCode', 'NewBusinessExpiryDate', 'RenewalExpiryDate', 'CreationDateTime'
    searchResultTableHeadersNamesOrder = [
        'RatebookName', 'RatebookID', 'RatebookVersion', 'RatebookStatusType', 'NewBusinessEffectiveDate',  'RenewalEffectiveDate',
        ]
    fieldsDict = {x.name: x for x in RatebookMetadata._meta.fields}
    searchResultTableHeaders = [fieldsDict[field] for field in searchResultTableHeadersNamesOrder]

    form = searchCriteriaForm()

    # save the Form data to session before validation and main action
    if request.method == 'POST':
        ratebook_details = request.POST.copy()
        request.session['PreviousSearchCriteria'] = ratebook_details
        form = searchCriteriaForm(ratebook_details)

    # Required for GET requests
    if request.method == 'GET' and request.session.get('PreviousSearchCriteria'):
        form = searchCriteriaForm(
            request.session['PreviousSearchCriteria']
        )

    # Mostly the form will only be valid if it is a POST request
    if form.is_valid():
        # save the cleaned Form data to session
        form_data = form.cleaned_data

        # check for existing template/Ratebook in production and if found show that it already exists.
        identityDetails = extractIdentityDetails(form_data)
        identityDetails.pop('Carrier')
        searchResults = RatebookMetadata.objects.filter(
                **identityDetails).order_by('-RatebookID')
        if 'Create a new Ratebook/Template' == request.POST.get('submit'):
            return redirect('ratemanager:projectIdAndDateInput')
        if searchResults.count() > 0 or request.POST.get('submit') == 'Search':

            # check for matching drafts if found show the draft.
            if searchResults.filter(RatebookStatusType='Draft').count() > 0:
                messages.add_message(
                    request, messages.INFO, RATE_MANAGER['MES_0001'])
            elif searchResults.count() > 0:
                messages.add_message(
                    request, messages.INFO, RATE_MANAGER['MES_0002'])
            else:
                messages.add_message(
                    request, messages.INFO, RATE_MANAGER['MES_0003'])

            searchResults.order_by(
                'RatebookID', 'RatebookStatusType', '-RatebookVersion'
            )

    context = {
                'searchCriteriaForm': form,
                'searchResults': searchResults,
                'searchResultTableHeaders': searchResultTableHeaders,
                'ENVIRONMENT_HIERARCHY': ENVIRONMENT_HIERARCHY
                }
    return context


def validateUpload(extractedRBDetails):
    return True


def transformUsingTemplateData(rbID, df, exhibitName):
    pd.options.mode.copy_on_write = True

    template = RatebookTemplate.objects.filter(
        RatebookID=rbID, RatebookExhibit__Exhibit=exhibitName
        )

    # coverages = template.first().ExhibitCoverages.all()
    rateVariables = [var.DisplayName for var in template.first().ExhibitVariables.all()]

    newdf = pd.DataFrame()

    def handle_ranges(df_in):
        df_in[str(i)+'1'] = pd.NA
        df_in[str(i)+'2'] = pd.NA
        for index, val in df_in[i].items():
            has_separator, separator = find_separator(val)
            if has_separator:
                df_in.loc[index, str(
                    i)+'1'], df_in.loc[index, str(i)+'2'] = val.split(separator)
            else:
                df_in.loc[index, str(i)+'1'] = val
        df_in.drop([i], axis=1, inplace=True)
        rateVariables.remove(i)
        rateVariables.extend([str(i)+'1', str(i)+'2'])

    df = df.melt(id_vars=rateVariables)
    df.rename(columns={'variable': 'Coverage',
                       'value': 'Factor'}, inplace=True)
    for i in df.columns:
        if check_range(df[i]):
            handle_ranges(df[i])
    newdf = create_rate_vars_cols(df, rateVariables)
    newdf['Exhibit'] = exhibitName

    return newdf


def transformRBTemplate(rbID=None, xl_url=None, df_sheets=None):
    """Function to transform the excel file to required form
    to be able to load into postgres database."""

    msgs = []
    if xl_url:
        df_sheets = pd.read_excel(xl_url, sheet_name=None)

    df_dict = dict()
    df_out = pd.DataFrame(columns=["Exhibit", "Coverage", "Factor"])

    for sheet_name, df in df_sheets.items():
        if sheet_name not in excludeList:
            df_dict[sheet_name] = transformUsingTemplateData(
                rbID=rbID, exhibitName=sheet_name, df=df)
            df_dict[sheet_name]['TableCategory'] = get_table_category(
                sheet_name)
            df_dict[sheet_name] = df_dict[sheet_name].astype(object)
            try:
                df_out = df_out.merge(df_dict[sheet_name], how='outer')
            except Exception as err:
                print("Unable to transform {} table due to {}".format(
                    sheet_name, err))

    numRatingVars = configs.NO_OF_RATING_VARIABLES
    for i in range(1, numRatingVars+1):
        if 'RatingVarName' + str(i) not in df_out.columns:
            df_out['RatingVarName' + str(i)] = None
            df_out['RatingVarValue' + str(i)] = None

    df_out = df_out[df_out.columns.sort_values().to_list()]
    df_out = df_out.astype(str).astype(object).where(pd.notnull(df_out), None)

    for i in set(df_sheets.keys()) - set(df_out["Exhibit"].unique()):
        if i not in excludeList:
            msgs.append("Unable to transform {} table.".format(i))
    # map_covs_and_vars(df_out)
    # df_out.to_excel('./uploads/check_transform.xlsx')
    return df_out, msgs
